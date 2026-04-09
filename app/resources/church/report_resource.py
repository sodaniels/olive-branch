# resources/church/report_resource.py

import time
import json
import csv
import io
from datetime import datetime
from flask import g, request, Response
from flask.views import MethodView
from flask_smorest import Blueprint
from pymongo.errors import PyMongoError

from ..doseal.admin.admin_business_resource import token_required
from ...models.church.report_model import AuditLog, ReportGenerator
from ...models.church.branch_model import Branch
from ...schemas.church.report_schema import (
    ReportGenerateQuerySchema, ReportExportQuerySchema, ReportAvailableQuerySchema,
    AuditLogQuerySchema, AuditLogCreateSchema,
)
from ...utils.json_response import prepared_response
from ...utils.helpers import make_log_tag, _resolve_business_id
from ...utils.logger import Log

blp_report = Blueprint("reports", __name__, description="Reporting, analytics, audit logs, and exports")


def _validate_branch(branch_id, target_business_id, log_tag=None):
    branch = Branch.get_by_id(branch_id, target_business_id)
    if not branch:
        if log_tag:
            Log.info(f"{log_tag} branch not found: {branch_id}")
        return None
    return branch


# ════════════════════════════ GENERATE REPORT ════════════════════════════

@blp_report.route("/report/generate", methods=["GET"])
class ReportGenerateResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(
        summary="Generate a report (membership, attendance, giving, visitor, event, volunteer, communication)",
        description="Returns JSON report data. Use /report/export for CSV/Excel/PDF.",
        security=[{"Bearer": []}],
    )
    def get(self, qd):
        client_ip = request.remote_addr
        user_info = g.get("current_user", {}) or {}
        account_type = user_info.get("account_type")
        auth_user__id = str(user_info.get("_id"))
        auth_business_id = str(user_info.get("business_id"))
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        log_tag = make_log_tag("report_resource.py", "ReportGenerateResource", "get", client_ip, auth_user__id, account_type, auth_business_id, target_business_id)

        branch_id = qd.get("branch_id")
        if not _validate_branch(branch_id, target_business_id, log_tag):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")

        report_type = qd["report_type"]
        Log.info(f"{log_tag} generating report: {report_type}")
        start_time = time.time()

        try:
            data = ReportGenerator.generate(
                report_type, target_business_id, branch_id=branch_id,
                start_date=qd.get("start_date"), end_date=qd.get("end_date"),
                group_by=qd.get("group_by"), top_n=qd.get("top_n"),
            )
            duration = time.time() - start_time
            Log.info(f"{log_tag} report generated in {duration:.2f}s")

            # Log the export
            AuditLog.log(
                target_business_id, "View", "Reports",
                description=f"Generated {report_type} report",
                performed_by=auth_user__id, ip_address=client_ip,
                resource_type="Report", resource_id=report_type,
                branch_id=branch_id,
            )

            return prepared_response(True, "OK", f"Report: {report_type}.", data={
                "report_type": report_type,
                "branch_id": branch_id,
                "parameters": {"start_date": qd.get("start_date"), "end_date": qd.get("end_date"), "group_by": qd.get("group_by"), "top_n": qd.get("top_n")},
                "data": data,
                "generated_at": datetime.utcnow().isoformat(),
                "generation_time_seconds": round(duration, 2),
            })
        except Exception as e:
            Log.error(f"{log_tag} error: {e}")
            return prepared_response(False, "INTERNAL_SERVER_ERROR", "Failed to generate report.", errors=[str(e)])


# ════════════════════════════ EXPORT REPORT ════════════════════════════

@blp_report.route("/report/export", methods=["GET"])
class ReportExportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportExportQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(
        summary="Export a report as CSV, Excel, or PDF",
        description="Returns downloadable file. Supported formats: csv, excel, pdf, json.",
        security=[{"Bearer": []}],
    )
    def get(self, qd):
        client_ip = request.remote_addr
        user_info = g.get("current_user", {}) or {}
        account_type = user_info.get("account_type")
        auth_user__id = str(user_info.get("_id"))
        auth_business_id = str(user_info.get("business_id"))
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        log_tag = make_log_tag("report_resource.py", "ReportExportResource", "get", client_ip, auth_user__id, account_type, auth_business_id, target_business_id)

        branch_id = qd.get("branch_id")
        if not _validate_branch(branch_id, target_business_id, log_tag):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")

        report_type = qd["report_type"]
        export_format = qd["format"]

        Log.info(f"{log_tag} exporting {report_type} as {export_format}")

        try:
            data = ReportGenerator.generate(
                report_type, target_business_id, branch_id=branch_id,
                start_date=qd.get("start_date"), end_date=qd.get("end_date"),
                group_by=qd.get("group_by"), top_n=qd.get("top_n"),
            )

            if data.get("error"):
                return prepared_response(False, "BAD_REQUEST", data["error"])

            filename = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"

            # Log the export
            AuditLog.log(
                target_business_id, "Export", "Reports",
                description=f"Exported {report_type} as {export_format}",
                performed_by=auth_user__id, ip_address=client_ip,
                resource_type="Report", resource_id=report_type,
                branch_id=branch_id,
            )

            if export_format == "json":
                return Response(
                    json.dumps(data, indent=2, default=str),
                    mimetype="application/json",
                    headers={"Content-Disposition": f"attachment; filename={filename}.json"},
                )

            if export_format == "csv":
                rows = _flatten_report_to_rows(data, report_type)
                output = io.StringIO()
                if rows:
                    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
                    writer.writeheader()
                    writer.writerows(rows)
                return Response(
                    output.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
                )

            if export_format == "excel":
                try:
                    import openpyxl
                    from io import BytesIO
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = report_type.replace("_", " ").title()

                    rows = _flatten_report_to_rows(data, report_type)
                    if rows:
                        # Header
                        headers = list(rows[0].keys())
                        for col, h in enumerate(headers, 1):
                            cell = ws.cell(row=1, column=col, value=h)
                            cell.font = openpyxl.styles.Font(bold=True)
                        # Data
                        for r_idx, row in enumerate(rows, 2):
                            for c_idx, h in enumerate(headers, 1):
                                ws.cell(row=r_idx, column=c_idx, value=row.get(h))
                        # Auto-width
                        for col in ws.columns:
                            max_len = max(len(str(cell.value or "")) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)

                    return Response(
                        buf.getvalue(),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
                    )
                except ImportError:
                    return prepared_response(False, "BAD_REQUEST", "openpyxl not installed. Use pip install openpyxl.")

            if export_format == "pdf":
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib import colors
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet
                    from io import BytesIO

                    buf = BytesIO()
                    doc = SimpleDocTemplate(buf, pagesize=A4)
                    styles = getSampleStyleSheet()
                    elements = []

                    # Title
                    elements.append(Paragraph(report_type.replace("_", " ").title(), styles["Heading1"]))
                    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
                    elements.append(Spacer(1, 12))

                    rows = _flatten_report_to_rows(data, report_type)
                    if rows:
                        headers = list(rows[0].keys())
                        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in rows]
                        t = Table(table_data)
                        t.setStyle(TableStyle([
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                        ]))
                        elements.append(t)
                    else:
                        elements.append(Paragraph("No data available for this report.", styles["Normal"]))

                    doc.build(elements)
                    buf.seek(0)

                    return Response(
                        buf.getvalue(),
                        mimetype="application/pdf",
                        headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
                    )
                except ImportError:
                    return prepared_response(False, "BAD_REQUEST", "reportlab not installed. Use pip install reportlab.")

            return prepared_response(False, "BAD_REQUEST", f"Unsupported format: {export_format}")

        except Exception as e:
            Log.error(f"{log_tag} export error: {e}")
            return prepared_response(False, "INTERNAL_SERVER_ERROR", "Failed to export report.", errors=[str(e)])


# ════════════════════════════ AVAILABLE REPORTS ════════════════════════════

@blp_report.route("/reports/available", methods=["GET"])
class ReportAvailableResource(MethodView):
    @token_required
    @blp_report.arguments(ReportAvailableQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="List all available report types with descriptions", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info)
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")

        reports = [
            {"key": "membership_growth", "name": "Membership Growth", "description": "Monthly new members, totals, active vs archived", "category": "Membership", "supports_date_filter": True},
            {"key": "membership_demographics", "name": "Membership Demographics", "description": "Gender, marital status, membership status breakdown", "category": "Membership", "supports_date_filter": False},
            {"key": "membership_status", "name": "Membership Status Breakdown", "description": "Status distribution including archived", "category": "Membership", "supports_date_filter": False},
            {"key": "attendance_by_service", "name": "Attendance by Service", "description": "Per-service attendance with averages by type", "category": "Attendance", "supports_date_filter": True},
            {"key": "attendance_trends", "name": "Attendance Trends", "description": "Attendance over time grouped by week or month", "category": "Attendance", "supports_date_filter": True},
            {"key": "attendance_by_group", "name": "Attendance by Group", "description": "Group/ministry attendance with per-session averages", "category": "Attendance", "supports_date_filter": True},
            {"key": "visitor_report", "name": "Visitor Report", "description": "First-timers, returning visitors, conversion funnel, monthly trends", "category": "Visitors", "supports_date_filter": True},
            {"key": "giving_by_fund", "name": "Giving by Fund", "description": "Donation totals per fund including unallocated", "category": "Giving", "supports_date_filter": True},
            {"key": "giving_by_donor", "name": "Giving by Donor", "description": "Top donors ranked by total contributions", "category": "Giving", "supports_date_filter": True},
            {"key": "giving_by_period", "name": "Giving by Period", "description": "Donation totals by month/year with fees and net amounts", "category": "Giving", "supports_date_filter": True},
            {"key": "event_report", "name": "Event Report", "description": "Event registrations, attendance rates, and revenue", "category": "Events", "supports_date_filter": True},
            {"key": "volunteer_report", "name": "Volunteer Report", "description": "Volunteer scheduling, RSVP rates, fulfilment by department", "category": "Volunteers", "supports_date_filter": True},
            {"key": "communication_report", "name": "Communication Report", "description": "Message delivery, open rates, click rates, bounce rates", "category": "Communication", "supports_date_filter": True},
            {"key": "audit_log", "name": "Audit Log", "description": "System access and action audit trail", "category": "System", "supports_date_filter": True},
        ]
        return prepared_response(True, "OK", f"{len(reports)} report types available.", data={"reports": reports, "count": len(reports), "export_formats": ["json", "csv", "excel", "pdf"]})


# ════════════════════════════ SHORTCUT ENDPOINTS ════════════════════════════

@blp_report.route("/reports/membership/growth", methods=["GET"])
class MembershipGrowthReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Membership growth report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.membership_growth(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Membership growth report.", data=data)

@blp_report.route("/reports/membership/demographics", methods=["GET"])
class MembershipDemographicsReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Membership demographics report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.membership_demographics(target_business_id, qd["branch_id"])
        return prepared_response(True, "OK", "Demographics report.", data=data)

@blp_report.route("/reports/attendance/trends", methods=["GET"])
class AttendanceTrendsReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Attendance trends report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.attendance_trends(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"), qd.get("group_by", "week"))
        return prepared_response(True, "OK", "Attendance trends.", data=data)

@blp_report.route("/reports/visitors", methods=["GET"])
class VisitorReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Visitor report with conversion funnel", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.visitor_report(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Visitor report.", data=data)

@blp_report.route("/reports/giving/by-fund", methods=["GET"])
class GivingByFundReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Giving by fund report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.giving_by_fund(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Giving by fund.", data=data)

@blp_report.route("/reports/giving/by-donor", methods=["GET"])
class GivingByDonorReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Giving by donor report (top N donors)", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.giving_by_donor(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"), qd.get("top_n", 50))
        return prepared_response(True, "OK", "Giving by donor.", data=data)

@blp_report.route("/reports/giving/by-period", methods=["GET"])
class GivingByPeriodReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Giving by period report (monthly/yearly totals with fees)", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.giving_by_period(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"), qd.get("group_by", "month"))
        return prepared_response(True, "OK", "Giving by period.", data=data)

@blp_report.route("/reports/events", methods=["GET"])
class EventReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Event performance report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.event_report(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Event report.", data=data)

@blp_report.route("/reports/volunteers", methods=["GET"])
class VolunteerReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Volunteer scheduling and fulfilment report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.volunteer_report(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Volunteer report.", data=data)

@blp_report.route("/reports/communication", methods=["GET"])
class CommunicationReportResource(MethodView):
    @token_required
    @blp_report.arguments(ReportGenerateQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Communication delivery and engagement report", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        data = ReportGenerator.communication_report(target_business_id, qd["branch_id"], qd.get("start_date"), qd.get("end_date"))
        return prepared_response(True, "OK", "Communication report.", data=data)


# ════════════════════════════ AUDIT LOG ════════════════════════════

@blp_report.route("/reports/audit-log", methods=["GET"])
class AuditLogResource(MethodView):
    @token_required
    @blp_report.arguments(AuditLogQuerySchema, location="query")
    @blp_report.response(200)
    @blp_report.doc(summary="Access and audit log with filters", security=[{"Bearer": []}])
    def get(self, qd):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info, qd.get("business_id"))
        if not _validate_branch(qd["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")
        r = AuditLog.get_all(
            target_business_id,
            action=qd.get("action"), module=qd.get("module"),
            performed_by=qd.get("performed_by"), branch_id=qd["branch_id"],
            resource_type=qd.get("resource_type"),
            start_date=qd.get("start_date"), end_date=qd.get("end_date"),
            page=qd.get("page", 1), per_page=qd.get("per_page", 50),
        )
        if not r.get("logs"):
            return prepared_response(False, "NOT_FOUND", "No audit logs found.")
        return prepared_response(True, "OK", "Audit logs.", data=r)


@blp_report.route("/reports/audit-log", methods=["POST"])
class AuditLogCreateResource(MethodView):
    @token_required
    @blp_report.arguments(AuditLogCreateSchema, location="json")
    @blp_report.response(201)
    @blp_report.doc(summary="Manually log an audit entry", security=[{"Bearer": []}])
    def post(self, json_data):
        user_info = g.get("current_user", {}) or {}
        target_business_id = _resolve_business_id(user_info)
        if not _validate_branch(json_data["branch_id"], target_business_id):
            return prepared_response(False, "NOT_FOUND", "Branch not found.")

        AuditLog.log(
            target_business_id,
            action=json_data["action"], module=json_data["module"],
            description=json_data.get("description"),
            performed_by=str(user_info.get("_id")),
            ip_address=request.remote_addr,
            resource_type=json_data.get("resource_type"),
            resource_id=json_data.get("resource_id"),
            branch_id=json_data["branch_id"],
            metadata=json_data.get("metadata"),
        )
        return prepared_response(True, "CREATED", "Audit log entry recorded.")


# ════════════════════════════ HELPER: FLATTEN REPORT DATA ════════════════════════════

def _flatten_report_to_rows(data, report_type):
    """Convert hierarchical report data into flat rows for CSV/Excel/PDF export."""
    rows = []

    # Membership
    if report_type == "membership_growth":
        for m in data.get("monthly_growth", []):
            rows.append({"Month": m["month"], "New Members": m["new_members"]})
        rows.append({"Month": "TOTAL", "New Members": data.get("period_new", 0)})

    elif report_type == "membership_demographics":
        for k, v in data.get("by_gender", {}).items():
            rows.append({"Category": "Gender", "Value": k, "Count": v})
        for k, v in data.get("by_membership_status", {}).items():
            rows.append({"Category": "Status", "Value": k, "Count": v})
        for k, v in data.get("by_marital_status", {}).items():
            rows.append({"Category": "Marital Status", "Value": k, "Count": v})

    elif report_type == "membership_status":
        for k, v in data.get("breakdown", {}).items():
            rows.append({"Status": k, "Count": v})

    # Attendance
    elif report_type == "attendance_by_service":
        for s in data.get("services", []):
            rows.append({"Date": s["date"], "Service Type": s["event_type"], "Attendance": s["attendance"]})

    elif report_type in ("attendance_trends",):
        for t in data.get("trends", []):
            rows.append({"Period": t["period"], "Attendance": t["attendance"]})

    elif report_type == "attendance_by_group":
        for g_item in data.get("groups", []):
            rows.append({"Group ID": g_item["group_id"], "Total Attendance": g_item["total_attendance"], "Sessions": g_item["sessions"], "Avg Per Session": g_item["avg_per_session"]})

    # Visitors
    elif report_type == "visitor_report":
        for m in data.get("monthly_visitors", []):
            rows.append({"Month": m["month"], "Visitors": m["visitors"]})
        rows.append({"Month": "SUMMARY", "Visitors": f"Total: {data.get('total_visitors', 0)}, Conversion: {data.get('conversion_rate', 0)}%"})

    # Giving
    elif report_type == "giving_by_fund":
        for f in data.get("by_fund", []):
            rows.append({"Fund": f["fund_name"], "Total": f["total"], "Donations": f["count"]})

    elif report_type == "giving_by_donor":
        for d_item in data.get("donors", []):
            rows.append({"Name": d_item["name"], "Total": d_item["total"], "Donations": d_item["count"], "First Gift": d_item.get("first_gift"), "Last Gift": d_item.get("last_gift")})

    elif report_type == "giving_by_period":
        for p in data.get("periods", []):
            rows.append({"Period": p["period"], "Total": p["total"], "Donations": p["count"], "Fees": p["fees"], "Net": p["net"]})

    # Events
    elif report_type == "event_report":
        for e in data.get("events", []):
            rows.append({"Event": e["name"], "Date": e["date"], "Registrations": e["registrations"], "Attendance": e["attendance"], "Rate": f"{e['attendance_rate']}%", "Revenue": e["revenue"]})

    # Volunteers
    elif report_type == "volunteer_report":
        for d_item in data.get("by_department", []):
            rows.append({"Department": d_item["department"], "Rosters": d_item["rosters"], "Total Assigned": d_item["total_assigned"]})
        rows.append({"Department": "SUMMARY", "Rosters": data.get("total_rosters", 0), "Total Assigned": f"Fulfilment: {data.get('fulfilment_rate', 0)}%"})

    # Communication
    elif report_type == "communication_report":
        rows.append({"Metric": "Total Messages", "Value": data.get("total_messages", 0)})
        rows.append({"Metric": "Sent", "Value": data.get("sent", 0)})
        rows.append({"Metric": "Delivered", "Value": data.get("delivered", 0)})
        rows.append({"Metric": "Failed", "Value": data.get("failed", 0)})
        rows.append({"Metric": "Delivery Rate", "Value": f"{data.get('delivery_rate', 0)}%"})
        rows.append({"Metric": "Open Rate", "Value": f"{data.get('open_rate', 0)}%"})
        rows.append({"Metric": "Click Rate", "Value": f"{data.get('click_rate', 0)}%"})

    else:
        # Generic fallback: dump as key-value pairs
        for k, v in data.items():
            if not isinstance(v, (dict, list)):
                rows.append({"Field": k, "Value": v})

    return rows
